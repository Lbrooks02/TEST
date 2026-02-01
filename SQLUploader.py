import os
import sys
import winsound
from datetime import datetime
import pyodbc                                           
from sqlalchemy import create_engine, text
import pandas as pd                                     
import numpy as np
from openpyxl import Workbook, load_workbook, styles
from ValidateData import Validate_Data                      #User defined function + file

def SQL_Connect(serverName: str, databaseName: str, connectionType: str = "sqlalchemy"):    #Create SQL Connection
    driverName = "ODBC Driver 17 for SQL Server"   
    Validate_Data("serverName", "databaseName")
    try:
        if connectionType.lower() == "sqlalchemy" :       
            sqlConnection = create_engine(f"mssql://@{serverName}/{databaseName}?driver={driverName}").connect()
        elif connectionType.lower() == "pyodbc" :
            sqlConnection = pyodbc.connect(f"DRIVER={{{driverName}}};SERVER={serverName};DATABASE={databaseName};Trusted_Connection=Yes;")
        else:
            Play_Sound("Error")
            sys.exit(">> Invalid SQL Connection Engine")
        print(f">> SQL connection successful: {sqlConnection}")
        return sqlConnection
    except Exception as e:
        Play_Sound("Error")
        print(e)
        sys.exit(">> SQL Connection failed!")

def File_Exists(filePath: str, printResult: bool = True) -> bool:                 #Check if file exists
    Validate_Data("filePath")
    fileExists = os.path.isfile(filePath) 
    if (not fileExists) and (printResult):
        Play_Sound("Error")
        print(f">> File '{filePath}' not found!")
    return fileExists

def Format_Datetime(dateTimeValue: datetime) -> str:    #Convert datetime value to string
    Validate_Data("dateTimeValue")
    try:
        formattedDateTime = datetime.strftime(dateTimeValue, format = '%Y-%m-%d %I:%M %p')
        return formattedDateTime                        #formats datetime to 12 hour AM/PM, returns as string
    except Exception as e:
        Play_Sound("Error")
        print(e)
        sys.exit(">> Error formatting datetime!")
    
def Get_Elapsed_Time(startTime: datetime, endTime: datetime) -> tuple:     #Return runtime metrics
    Validate_Data("startTime", "endTime")
    duration = (endTime - startTime).total_seconds()    # Get the total time in seconds
    days = divmod(duration, 86400)                      # Returns tuple (fullDays, remaining seconds)
    hours = divmod(days[1], 3600)                       # Uses remaining seconds from days[1], returns tuple (fullHours, remaining seconds)
    minutes = divmod(hours[1], 60)                      # Uses remaining seconds from hours[1], returns tuple (fullMinutes, remaining seconds)
    seconds = divmod(minutes[1], 1)                     # Uses remaining seconds from seconds[1], returns tuple (fullHours, 0 remaining seconds)
    stopWatch = f"{int(hours[0]):02d}:{int(minutes[0]):02d}:{int(seconds[0]):02d}"    #Format time as stopwatch (HH:MM:SS)
    timeString = f"{days[0]} days, {hours[0]} hours, {minutes[0]} minutes and {seconds[0]} seconds"
    elapsedTime = (stopWatch, timeString)
    return elapsedTime

def Log_Upload(tableName: str, recordCount: int, elapsedTime: str, uploadStartTime: datetime, uploadEndTime: datetime, sourceFile: str) -> None: #Export execution data to excel
    Validate_Data("tableName", "recordCount", "elapsedTime", "uploadStartTime", "uploadEndTime", "sourceFile")
    sheetName = "Upload_Results"
    workBookPath = os.path.join(os.path.expanduser("~"), "Documents", "SQL_Upload_Log.xlsx")  #Import Data is logged here
    headers = ['Table Name', 'Record Count', 'Runtime', 'Start Time','End Time','Source File']
    newLog = [tableName, recordCount, elapsedTime, Format_Datetime(uploadStartTime), Format_Datetime(uploadEndTime), sourceFile]
    if File_Exists(workBookPath, False):                       #Open workbook and go to main Sheet
        createFile = False
        workBook = load_workbook(workBookPath)
        if sheetName not in workBook.sheetnames:
            workBook.create_sheet(sheetName)
            workBook[sheetName].append(headers)
        workSheet = workBook[sheetName]    
    else:                                               #Create new workbook if one doesn't exist
        createFile = True
        workBook = Workbook()
        workSheet = workBook.active
        workSheet.title = sheetName
        workSheet.append(headers)
        for i in range(1, len(headers) + 1):          #Create/format header row
            workSheet.cell(row = 1, column = i).font = styles.Font(bold = True)         
            workSheet.cell(row = 1, column = i).border = styles.borders.Border(bottom = styles.borders.Side(style = 'thin'))
            workSheet.cell(row = 1, column = i).fill = styles.fills.PatternFill(patternType ='solid', fgColor = styles.colors.Color(rgb = 'BFBFBF'))
    workSheet.append(newLog)
    response = ""
    while True:                                     #Error handler if workbook is already open, loops this until closed
        try:
            workBook.save(workBookPath)             #Logs import data to Excel
            if createFile == True:
                print(f">> New file created at '{workBookPath}'")
            break
        except Exception as e:
            Play_Sound("Error")
            print(e)
            response = input(f">> SQL_Upload_Log is currently open. Close file at: '{workBookPath}' then hit the enter key to continue >> ")     
            if response.upper() == 'STOP':
                sys.exit(">> Cancel input command received")
    print (f">> Upload logged at '{workBookPath}'")

def Play_Sound(soundType: str) -> None:                     #Plays sound based on process success/failure
    Validate_Data("soundType")
    if soundType.lower() == "success":
        soundSource = r"C:\Windows\Media\Alarm03.wav"
    elif soundType.lower() == "error":
        soundSource = r"C:\Windows\Media\chord.wav"
    else:
        return None
    winsound.PlaySound(soundSource, winsound.SND_FILENAME | winsound.SND_NODEFAULT)

def SQL_Upload(sourceFile: str, serverName: str, databaseName: str, tableName: str, datasetName: str = "", skipRows: list = None, skipColumns: list = None, convertColumnTypes: dict = None, dateTimeFormat: str = None, delimiterCharacter: str = '|', encodingType: str = 'latin1') -> None: #Send data to SQL
    r"""
        Description:
            This process ingests data from Excel (.xls, .xlsx), flat files (.csv, .txt), 
            or Access databases (.accdb, .mdb) into a page-compressed SQL Server table. 
            Additionally, it appends execution metrics to a centralized log located within the user's Documents folder.
        Parameters:
            :sourceFile: String to store source file path
            :tableName: String to store SQL destination table name 
            :datasetName: *Optional* String to store datasheet/table name (Excel/Access)
            :skipRows: *Optional* List containing row indexes to skip (0 indexed, negative indexes are allowed)
            :skipColumns: *Optional* List containing column indexes to skip (0 indexed, negative indexes are allowed)
            :convertColumnTypes: *Optional* Dictionary containing {column:expectedDataType} pairs. Value must be in ("float", "str", "int", "datetime", "bool")
            :dateTimeFormat: *Optional* String to store datetime format so pandas can parse date strings into datetime values (Ex. "%d.%m.%Y")
            :delimiterCharacter: *Optional* String to store delimiter character (for .txt and .csv files)
            :encodingType: *Optional* String to store encoding style for special characters ("latin1" by default, can also use "utf8")
        Example Usage:
            parentFolder = "C:\Users\USERNAME\Documents"
            dataTypes = {
                'Column1': 'str',
                'Column2': 'int', 
                'Column3': 'bool', 
                'Column4': 'float', 
                'Column5': 'datetime', 
                }
            badRows = [0, 1]
            badCols = [0, 1, 2]
            SQL_Upload(fr"{parentFolder}\sourceFile.xlsx", "ServerName", "DBName", "SQLtable", convertColumnTypes=dataTypes, skipRows=badRows, skipColumns=badCols)
    """
    Validate_Data("sourceFile", "serverName", "databaseName", "tableName")
    #Set defaults (defaults in parameter list causes data leaks between function calls)
    skipRows = skipRows or []
    skipColumns = skipColumns or [] 
    convertColumnTypes = convertColumnTypes or {}
    if not dateTimeFormat:
       dateTimeFormat = "%d.%m.%Y"  #Default value
    if not File_Exists(sourceFile):
        sys.exit(1)
    sqlConnection = SQL_Connect(serverName, databaseName)
    try:
        try:
            uploadStartTime = datetime.now()                                #Setup SQL connection object
            print(f">> File reading started at: {Format_Datetime(uploadStartTime)}.")
            forceColumnType = bool(convertColumnTypes)                      #Returns True if user entered parameter
            specifyDatasetName = bool(datasetName.strip())                #Returns True if user entered parameter
            isExcelFile = bool(sourceFile.lower().endswith((".xls", ".xlsx")))
            isTextFile = bool(sourceFile.lower().endswith(".txt"))
            isCSVFile = bool(sourceFile.lower().endswith(".csv"))
            isACCDBFile = bool(sourceFile.lower().endswith((".accdb", ".mdb")))
            if isExcelFile: 
                fileExtension = os.path.splitext(sourceFile)[1].lower()
                if fileExtension == ".xls":
                    engine = "xlrd"
                elif fileExtension == ".xlsx":
                    engine = "openpyxl"
                else:
                    sys.exit(f">> Unsupported Excel format: {fileExtension}")
                if specifyDatasetName:                                #If a sheet name is specified
                    datasetName = datasetName.strip()
                    excelObject = pd.ExcelFile(sourceFile, engine=engine)
                    sheetList = excelObject.sheet_names
                    if datasetName in sheetList:
                        try:
                            df = pd.read_excel(sourceFile, sheet_name = datasetName, skiprows = [row for row in skipRows], engine=engine)  #Load excel into dataframe
                        except IndexError:
                            sys.exit(f">> One or more arguments supplied are out of range for file: {sourceFile}")
                    else:
                        sys.exit(f">> Sheet name does not exist in file: {sourceFile}")
                else:   
                    try:                                                    
                        df = pd.read_excel(sourceFile, skiprows = [row for row in skipRows], engine=engine)                          #Default to first sheet in workbook
                    except IndexError:
                        sys.exit(f">> One or more arguments supplied are out of range for file: {sourceFile}")
            elif (isTextFile) or (isCSVFile):
                try:
                    df = pd.read_csv(                                   #Update parameters based on the format of sourceFile
                        sourceFile, 
                        encoding = encodingType,                        #'latin1' interprets EU special letters/chars, default is 'utf8'
                        delimiter = delimiterCharacter,                 #Column separating character
                        header = 0,                                     #Identifies the column headers row number 
                        skiprows = [row for row in skipRows],           #Excludes rows from df
                        skipfooter = 0,                                 #Number of rows to exclude, starting from the bottom of df (update this to match dataset, rows>2 may be out of bounds)
                        engine = 'python'                               #Explicitly state engine to use skipfooter parameter
                    ) 
                except IndexError:
                    sys.exit(f">> One or more arguments supplied are out of range for file: {sourceFile}")
                for rsColumn in df.columns:                              
                    if df[rsColumn].dtypes == 'object':             #Points to specific column in dataframe 
                        df[rsColumn] = df[rsColumn].str.strip()     #Remove whitespace from records in stringType columns
                df.columns = df.columns.str.strip()                 #Remove whitespace from df column names
            elif isACCDBFile:
                accessDriverName = "Microsoft Access Driver (*.mdb, *.accdb)"
                accessDatabaseName = sourceFile
                try:
                    with pyodbc.connect(f"DRIVER={{{accessDriverName}}};DBQ={accessDatabaseName}") as accessConnection:
                        print(f">> Access connection successful: {accessConnection}")
                        cursor = accessConnection.cursor()
                        tableList = [row.table_name for row in cursor.tables(tableType="TABLE")]
                        if len(tableList) == 0:
                            sys.exit(f">> No tables exist in the db: {accessDatabaseName}")
                        if specifyDatasetName == True: 
                            tableName = datasetName.strip().replace("[", "").replace("]", "")
                            if tableName not in tableList:
                                sys.exit(f">> Sheet does not exist in the db: {accessDatabaseName}")
                        else:
                            sys.exit(f">> No table supplied from db: {accessDatabaseName}")
                        try:
                            df = pd.read_sql(f"SELECT * FROM [{tableName}]", accessConnection)
                        except Exception as e:
                            Play_Sound("Error")
                            print(e)
                            sys.exit(f">> Failed to read table {tableName} from db: {accessDatabaseName}")
                except Exception as e:
                    Play_Sound("Error")
                    print(e)
                    sys.exit(">> Access connection failed!")
            else:
                sys.exit(f">> Unsupported file type: {sourceFile}")
            try:
                df = df.drop(df.columns[skipColumns], axis = 1)         #Drops columns from df
            except IndexError:
                sys.exit(f">> One or more arguments supplied are out of range for file: {sourceFile}")
            alterColumnSQL = []
            if forceColumnType == True:       
                for columnName in convertColumnTypes:
                    for rsColumn in df.columns:
                        try:
                            if columnName == rsColumn:
                                convertDatatype = convertColumnTypes[columnName]  #Retrieves dataType from convertColumnTypes
                                match convertDatatype:
                                    case 'datetime':                 #Format column as datetime (May need to remove the excel condition, if date columns are auto-read as int64 in .txt files )
                                        if isExcelFile:
                                            if (df[rsColumn].dtypes == 'int64') or (df[rsColumn].dtypes == 'Int64'): 
                                                df[rsColumn] = pd.to_datetime('1899-12-30') + pd.to_timedelta(df[rsColumn], 'D') #Dates in excel are read as integers
                                            elif df[rsColumn].dtypes == 'float64': 
                                                df[rsColumn] = pd.to_datetime('1899-12-30') + pd.to_timedelta((np.floor(df[rsColumn])).astype('Int64'), 'D') #Dates in excel are read as integers
                                        else:
                                            df[rsColumn] = pd.to_datetime(df[rsColumn], format = dateTimeFormat).astype('datetime64[ns]') #Convert string column to datetime (May need to adjust format parameter to match input file)
                                    case 'float':                  #Format column as float
                                        if df[rsColumn].dtypes == 'object':
                                            df[rsColumn] = df[rsColumn].str.replace('$','').str.replace(',','')
                                        df[rsColumn] = pd.to_numeric(df[rsColumn]).astype('float64')
                                    case 'int':                    #Format column as integer
                                        if df[rsColumn].dtypes == 'object':
                                            df[rsColumn] = df[rsColumn].str.replace('$','', regex=False).str.replace(',','', regex=False)
                                        df[rsColumn] = (np.floor(pd.to_numeric(df[rsColumn]))).astype('Int64')
                                    case 'str':                    #Format column as string
                                        df[rsColumn] = df[rsColumn].astype('str')
                                    case 'bool':                   #Format column as boolean
                                        if df[rsColumn].dtypes == 'object':
                                            df[rsColumn] = df[rsColumn].str.upper().replace({'TRUE':True, 'FALSE':False})
                                        df[rsColumn] = df[rsColumn].astype('boolean')
                        except Exception as e:
                            Play_Sound("Error")
                            print(e)
                            sys.exit(f">> Error formatting column '{rsColumn}'!")
            for col in df.columns: 
                dataType = df[col].dtypes
                if dataType == "object":    
                    df.loc[df[col].isin(['', 'nan']), col] = np.nan   #Null/blank handling - Update values to null if blank string or 'nan'
                    mSQL = f"ALTER TABLE dbo.[{tableName}] ALTER COLUMN [{col}] VARCHAR(255)"
                elif dataType in ("Int64", "int64"):
                    mSQL = f"ALTER TABLE dbo.[{tableName}] ALTER COLUMN [{col}] INTEGER"
                elif dataType in ("bool", "boolean"):
                    mSQL = f"ALTER TABLE dbo.[{tableName}] ALTER COLUMN [{col}] BIT"
                else:
                    continue
                alterColumnSQL.append(mSQL)
        except Exception as e:
            Play_Sound("Error")
            print(e)
            sys.exit(">> Error reading file!")
        print (f">> Dataframe loaded [{len(df.index)} rows x {len(df.columns)} columns]")
        print("\n")
        df.info()   #print class info for DF
        #print(f"\n{df}\n")  #print full df
        print(f"\n{df.head()}\n") #prints limited snippet of dataframe
        recordCount = len(df.index)
        print(f">> Upload started at: {Format_Datetime(datetime.now())}")
        try:
            mSQL = f"DROP TABLE IF EXISTS dbo.[{tableName}]"
            sqlConnection.execute(text(mSQL))
            sqlConnection.commit()
            df.to_sql(tableName, sqlConnection, schema = 'dbo', if_exists = 'replace', index = False)   #Send to SQL
            try:
                mSQL = f"ALTER TABLE dbo.[{tableName}] REBUILD PARTITION = ALL WITH (DATA_COMPRESSION = PAGE)"
                sqlConnection.execute(text(mSQL))
                sqlConnection.commit()
                print(f">> [{tableName}] successfully compressed.")
            except Exception as e:
                Play_Sound("Error")
                print(e)
                sys.exit(f">> Error compressing [{tableName}]")
        except Exception as e:
            Play_Sound("Error")
            print(e)
            sys.exit(">> SQL import failed!")
        for mSQL in alterColumnSQL:
            try:
                sqlConnection.execute(text(mSQL))   #bigint -> int conversion may fail if number is larger than '2,147,483,647'
                sqlConnection.commit()
            except Exception as e:
                Play_Sound("Error")
                print(e)
                print(f">> Error running SQL: '{mSQL}'")
                continue    
        uploadEndTime = datetime.now()
        sqlConnection.close()
        del df 
        print(f">> Upload finished at: {Format_Datetime(uploadEndTime)}")
        elapsedTime = Get_Elapsed_Time(uploadStartTime, uploadEndTime)                                  #Tracks session time
        Log_Upload(tableName, recordCount, elapsedTime[0], uploadStartTime, uploadEndTime, sourceFile)  #Append upload session data to log file
        Play_Sound("Success")
        print(f"***** [{tableName}] upload completed in: {elapsedTime[1]}*****", end = "\n\n")
    except KeyboardInterrupt:       #Allows user to interrupt the program with Ctrl+C
        sys.exit(">> Cancel input command received")
        

